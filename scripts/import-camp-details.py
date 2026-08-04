"""Replace all seeded/mock camp data with the real camps from Camp Details.xlsx.

The workbook is one sheet per district with columns:
    ROW | CAMP NAME | LOCALITY | TALUK | DISTRICT | PERSON OF CONTACT |
    CONTACT NUMBER | INFORMATION SOURCE
Pathanamthitta carries two extra columns (Name, Contact).

    uv run --project api python scripts/import-camp-details.py [--dry-run]

DESTRUCTIVE: removes every existing camp and the synthetic camp_needs rows.
"""

from __future__ import annotations

import argparse
import asyncio
import pathlib
import re
import sys

import asyncpg
import openpyxl

ROOT = pathlib.Path(__file__).resolve().parent.parent
WORKBOOK = ROOT / "Camp Details.xlsx"

# Free-text columns get this when the sheet leaves them blank, so the UI shows
# an explicit "we don't know" rather than an ambiguous gap.
UNFILLED = "Unfilled"

# Sheet spellings vary from the districts table ("Trissur" vs "Thrissur").
DISTRICT_CODES: dict[str, str] = {
    "thiruvananthapuram": "TVM", "tvm": "TVM", "trivandrum": "TVM",
    "kollam": "KLM",
    "pathanamthitta": "PTA",
    "alappuzha": "ALP", "alleppey": "ALP",
    "kottayam": "KTM",
    "idukki": "IDK",
    "ernakulam": "EKM", "kochi": "EKM",
    "thrissur": "TSR", "trissur": "TSR",
    "palakkad": "PKD",
    "malappuram": "MPM",
    "kozhikode": "KKD", "calicut": "KKD",
    "wayanad": "WYD",
    "kannur": "KNR",
    "kasaragod": "KSD",
}

PHONE_IN_TEXT = re.compile(r"(?:\+?91[\s-]*)?(\d[\d\s-]{8,13}\d)")


def clean(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def normalise_phone(raw: object) -> str | None:
    """Return +91XXXXXXXXXX, or None.

    None rather than "Unfilled" on purpose: the UI turns this column into a
    tel: link, and a link that dials the word "unfilled" is worse than the
    honest "No phone number reported for this camp." the UI already shows.
    """
    text = clean(raw)
    if not text:
        return None
    digits = re.sub(r"\D", "", text)
    if len(digits) < 10:
        return None
    ten = digits[-10:]
    return f"+91{ten}" if re.fullmatch(r"[6-9]\d{9}", ten) else None


def split_source(raw: object) -> tuple[str | None, str | None]:
    """"ChandraBose +91 92072 92007" -> ("ChandraBose", "+919207292007")."""
    text = clean(raw)
    if not text:
        return None, None

    match = PHONE_IN_TEXT.search(text)
    if not match:
        return text, None

    phone = normalise_phone(match.group(0))
    name = text[: match.start()].strip(" ,-") or None
    return name, phone


def read_rows() -> list[dict[str, object]]:
    workbook = openpyxl.load_workbook(WORKBOOK, data_only=True)
    out: list[dict[str, object]] = []
    skipped: list[str] = []

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue

        header = [str(h).strip().upper() if h else "" for h in rows[0]]

        def col(name: str, header: list[str] = header) -> int | None:
            # Default-arg binding: without it this closure would read whatever
            # `header` happens to be on a later sheet.
            return header.index(name) if name in header else None

        i_name, i_local = col("CAMP NAME"), col("LOCALITY")
        i_taluk, i_district = col("TALUK"), col("DISTRICT")
        i_person, i_phone = col("PERSON OF CONTACT"), col("CONTACT NUMBER")
        i_source = col("INFORMATION SOURCE")
        # Pathanamthitta's extra pair.
        i_alt_name = header.index("NAME") if "NAME" in header else None
        i_alt_phone = header.index("CONTACT") if "CONTACT" in header else None

        def cell(row: tuple, index: int | None) -> object:
            return row[index] if index is not None and index < len(row) else None

        for row in rows[1:]:
            name = clean(cell(row, i_name))
            if not name:
                continue  # blank filler row

            district_raw = clean(cell(row, i_district)) or sheet_name
            code = DISTRICT_CODES.get(district_raw.strip().lower())
            if not code:
                skipped.append(f"{sheet_name}: unknown district {district_raw!r} for {name!r}")
                continue

            source_text = clean(cell(row, i_source))
            source_name, source_phone = split_source(source_text)

            contact_name = (
                clean(cell(row, i_person))
                or clean(cell(row, i_alt_name))
                or source_name
                or UNFILLED
            )
            phone = (
                normalise_phone(cell(row, i_phone))
                or normalise_phone(cell(row, i_alt_phone))
                or source_phone
            )

            locality = clean(cell(row, i_local)) or UNFILLED

            out.append(
                {
                    "name": name,
                    "district_code": code,
                    "taluk": clean(cell(row, i_taluk)) or UNFILLED,
                    "village_or_locality": locality,
                    # The sheet has no local-body column; locality is the closest
                    # real signal, and the type is unknowable from this source.
                    "lsg_name": locality,
                    "lsg_type": "panchayat",
                    "camp_incharge_name": contact_name,
                    "camp_phone_primary": phone,
                    "source_label": source_text or UNFILLED,
                }
            )

    for line in skipped:
        print(f"  SKIPPED {line}")
    return out


async def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    camps = read_rows()
    print(f"\nparsed {len(camps)} camps from {WORKBOOK.name}")

    by_district: dict[str, int] = {}
    for camp in camps:
        by_district[str(camp["district_code"])] = by_district.get(str(camp["district_code"]), 0) + 1
    print("  by district:", dict(sorted(by_district.items())))
    print(f"  with a phone number: {sum(1 for c in camps if c['camp_phone_primary'])}")

    if args.dry_run:
        print("\n-- dry run, nothing written --")
        for camp in camps[:5]:
            print("  ", camp)
        return

    if not camps:
        sys.exit("refusing to wipe the database with nothing to import")

    env = (ROOT / "api" / ".env").read_text(encoding="utf-8")
    match = re.search(r"^SUPABASE_DB_URL_DIRECT=(.+)$", env, re.M)
    if not match:
        sys.exit("SUPABASE_DB_URL_DIRECT missing from api/.env")

    connection = await asyncpg.connect(match.group(1).strip(), statement_cache_size=0)
    try:
        async with connection.transaction():
            # Remove the seeded/synthetic data. Children first.
            print("\nclearing existing data")
            for table in (
                "need_pledges", "camp_needs", "report_images", "reports",
                "camp_checkins", "image_flags", "camp_sources", "audit_log",
            ):
                await connection.execute(f"DELETE FROM {table}")
            await connection.execute("DELETE FROM camps")
            await connection.execute("DELETE FROM sources")

            source_id = await connection.fetchval(
                """
                INSERT INTO sources (type, label, reliability_note)
                VALUES ('internal_volunteer', 'Camp Details volunteer sheet',
                        'Compiled by on-ground volunteers; contacts vary in completeness.')
                RETURNING id
                """
            )

            for camp in camps:
                camp_id = await connection.fetchval(
                    """
                    INSERT INTO camps (
                        name, district_code, taluk, lsg_type, lsg_name,
                        village_or_locality, camp_incharge_name, camp_phone_primary,
                        verification_state, status, urgency, report_count,
                        status_last_confirmed_at, amenities
                    )
                    VALUES ($1,$2,$3,$4::lsg_type,$5,$6,$7,$8,
                            'unverified','active','normal',1, now(), '{}')
                    RETURNING id
                    """,
                    camp["name"], camp["district_code"], camp["taluk"], camp["lsg_type"],
                    camp["lsg_name"], camp["village_or_locality"],
                    camp["camp_incharge_name"], camp["camp_phone_primary"],
                )
                # Provenance: every camp links back to where it came from.
                await connection.execute(
                    "INSERT INTO camp_sources (camp_id, source_id) VALUES ($1, $2)",
                    camp_id, source_id,
                )

        total = await connection.fetchval("SELECT count(*) FROM camps")
        print(f"\nimported. camps now: {total}")
        rows = await connection.fetch(
            "SELECT district_code, count(*) AS n FROM camps GROUP BY 1 ORDER BY 1"
        )
        for row in rows:
            print(f"  {row['district_code']}  {row['n']}")
    finally:
        await connection.close()


if __name__ == "__main__":
    asyncio.run(main())
